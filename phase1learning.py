import pyttsx3
import speech_recognition as sr
import pyaudio
import smtplib
import os
from email.message import EmailMessage
import imghdr
import fileinput
import matplotlib.pyplot as plt
import numpy as np
from sklearn.linear_model import LogisticRegression
import pandas as pd
from openpyxl import load_workbook
import openpyxl

engine = pyttsx3.init('sapi5')

voices = engine.getProperty('voices')
engine.setProperty('voices', voices[0].id)


def speak(audio):
    engine.say(audio)
    engine.runAndWait()

if __name__ == "__main__":
    pass

def takecommand():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        print("Listening")
        r.pause_threshold = 1
        audio = r.listen(source)

    try:
        print("Recognizing")
        query = r.recognize_google(audio,language="en-in")
        print(query)

    except Exception as e:
        print("Speak again")
        return "None"

    return query


if __name__ == "__main__":
    answer = np.array([])
    while True:
        query = takecommand().lower()

        if "send a mail" in query:
            speak("Initiating your gmail account")
            speak("Are you preparing for sample mail?")
        if "yes" in query:
            speak("Subject please")
        if "leave application" in query:
            Receiver = input("Enter the Receiver name")
            Date = input("Date")
            Time = input("No. of days")
            Sender = input("Your Name please")
            Coworker = input("Coworker name")

            contacts = ['nagpurenishad4@gmail.com']

            replace_texts = {"Receiver":Receiver,
                            "Date":Date,
                            "din":Time,
                            "Name":Sender,"Collegue":Coworker}

            for line in fileinput.input("C:\\Users\\admin\\Desktop\\work\\testing.txt",inplace=True):
                for search_text in replace_texts:
                    replace_text = replace_texts[search_text]
                    line = line.replace(search_text,replace_text)
                print(line,end=' ')

            msg = EmailMessage()
            msg['Subject'] = "Leave Application"
            msg['From'] = "nagpurepramod877@gmail.com"
            msg['To'] = contacts
            msg.set_content("Please find a attachment below")

            files = ["C:\\Users\\admin\\Desktop\\work\\testing.txt"]

                                                
            for file in files:
                with open(file,'rb') as f:
                    file_data = f.read()
                    file_name = f.name
                msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)

            with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
                smtp.ehlo()
                smtp.starttls()
                smtp.ehlo()

                smtp.login("nagpurepramod877@gmail.com","Nishad@P3")

                smtp.send_message(msg)

            place_texts ={Receiver:"Receiver",Date:"Date",Time:"din",Sender:"Name",Coworker:"Collegue"}

            for line in fileinput.input("C:\\Users\\admin\\Desktop\\work\\testing.txt",inplace=True):
                for search_text in place_texts:
                    place_text = place_texts[search_text]
                    line = line.replace(search_text,place_text)
                print(line,end=' ')


        if 'coffee' in query:
            l = input('When was your last meal (in hrs)')
            f = input ('when are you going to have your future meal (in hrs.)')
            wb = load_workbook('C:\\Users\\admin\\Desktop\\logisticmodel.xlsx')
            ws =wb.worksheets[0]
            ws['A2'] = l
            ws['B2'] = f
            wb.save('C:\\Users\\admin\\Desktop\\logisticmodel.xlsx')
            df_2 = pd.read_excel('C:\\Users\\admin\\Desktop\\logisticmodel.xlsx')
            df = pd.read_excel("C:\\Users\\admin\\Desktop\\work\\cafesample.xlsx")
            train = df[0:]
            train_y = train.pop('coffee_size')
            train_x_1 = [df['future_meal'], df['last_meal']]
            headers = ['future_meal','last_meal']
            train = pd.concat(train_x_1, axis=1,keys = headers)
            train_x = np.asarray(train)
            train_y = np.asarray(train_y)
            test = np.asarray(df_2)
            log_reg = LogisticRegression()
            log_reg.fit(train_x,train_y)
            test = np.asarray(df_2)
            prediction = log_reg.predict(test)
            if prediction[0]==0:
                speak("I say you should go with large size cup")
            if prediction[0]==1:
                speak("I guess you should go with medium size cup")
            if prediction[0]==2:
                speak("I recommend you a small size cup")

            speak("Sir can i ask you what size coffee will you prefer")
            while True:
                query = takecommand().lower()
                if 'large' in query:
                    speak('ok')
                if 'medium' in query:
                    speak('ok')
                if 'small' in query:
                    speak('ok')
                break
            data = np.asarray(df)
            data_2 = data.flatten()
            feature = [0,f,l]
            feature_array = np.asarray(feature)
            new_data = np.append(data_2,feature)
            datin = new_data.reshape(-1,3)
            data_frame = pd.DataFrame(datin, columns=['coffee_size','future_meal','last_meal'])
            data_frame.to_excel("C:\\Users\\admin\\Desktop\\work\\cafesample.xlsx", index=False)


                        
            





            












 
  